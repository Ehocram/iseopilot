"""
main.py — Applicazione FastAPI (ISEOPilot) — Incremento 2.

Autenticazione locale gestita dall'admin (sessioni con cookie firmato).
Tutte le pagine sono protette dal login, tranne /login e /healthz.
"""
from __future__ import annotations

import json
import re
import os
import threading
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from starlette.middleware.sessions import SessionMiddleware
from starlette.background import BackgroundTask

from . import diag
diag.install()

from . import auth, connectors, docgen, i18n, knowledge, memory, store
from .connectors import (DEF_OD_CLIENT_ID, DEF_OD_TENANT_ID, DEF_DYN_CLIENT_ID,
                         DEF_DYN_TENANT_ID, DEF_DYN_RESOURCE_URL,
                         DEF_PBI_CLIENT_ID, DEF_PBI_TENANT_ID)
from .orchestrator import TONES, LANG_INSTR, stream_reply

BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

app = FastAPI(title="ISEOPilot")

# Sessione firmata. Richiede APP_SECRET_KEY (fail loudly se assente).
_secret = os.environ.get("APP_SECRET_KEY", "").strip()
if not _secret:
    raise RuntimeError("APP_SECRET_KEY non impostata: necessaria per firmare le sessioni.")


@app.middleware("http")
async def _no_html_cache(request, call_next):
    """Le pagine HTML non si cacheano MAI (browser e App Proxy): l'HTML vecchio
    referenzia asset vecchi col loro ?v= vecchio, e il cache-busting non può
    bustare sé stesso. Classe di incidenti: 'funziona solo dopo hard-refresh'."""
    response = await call_next(request)
    ctype = response.headers.get("content-type", "")
    if ctype.startswith("text/html"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
    return response


app.add_middleware(
    SessionMiddleware,
    secret_key=_secret,
    same_site="lax",
    https_only=os.environ.get("SESSION_HTTPS_ONLY", "0") == "1",
    max_age=60 * 60 * 8,  # 8 ore
)

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


# Versione degli asset statici: cambia quando i file cambiano (deploy), così
# il browser NON serve dalla cache un chat.js vecchio dopo un aggiornamento.
try:
    _static_dir = Path(__file__).resolve().parent.parent / "static"
    ASSET_V = str(int(max(f.stat().st_mtime for f in _static_dir.glob("*") if f.is_file())))
except Exception:
    ASSET_V = "1"


@app.on_event("startup")
def _startup():
    store.init_db()
    auth.bootstrap_admin()
    # Come l'app desktop: indicizza le cartelle configurate all'avvio, in
    # background (anche le share di rete), così sono subito ricercabili.
    knowledge.reindex_all_async()
    # Warm-up dell'indice semantico Dynamics (se il catalogo esiste): senza,
    # il PRIMO utente del processo pagava 70-120s di costruzione embeddings.
    def _warm_dyn():
        try:
            from .engines import dynamics_search as _ds
            _ds.warm_semantic_index({})
        except Exception:
            pass
    threading.Thread(target=_warm_dyn, daemon=True).start()
    threading.Thread(target=_attach_purge_old, daemon=True).start()


def _ui_lang(request: Request, user: dict | None = None) -> str:
    """Lingua interfaccia: preferenza utente se loggato, altrimenti sessione, default 'it'."""
    if user:
        return i18n.normalize(store.get_user_setting(user["username"], "ui_lang", "it"))
    return i18n.normalize(request.session.get("ui_lang", "it"))


def _i18n_ctx(request: Request, user: dict | None = None) -> dict:
    lang = _ui_lang(request, user)
    return {
        "t": (lambda key: i18n.t(key, lang)),
        "ui_lang": lang,
        "next_path": request.url.path,
        "asset_v": ASSET_V,
    }


def _ctx(request: Request, user: dict, **extra) -> dict:
    try:
        from . import dubstudio as _dub
        _dub_nav = _dub.enabled() and _dub.user_allowed(user["username"])
    except Exception:
        _dub_nav = False
    base = {"user": user["username"], "is_admin": bool(user["is_admin"]),
            "department": user.get("department") or "—",
            "dub_nav": _dub_nav}
    base.update(_i18n_ctx(request, user))
    base.update(extra)
    return base


# Budget del contesto allegati: con Claude (200k token) i vecchi tagli a
# 12-14k caratteri TOTALI facevano sparire i file oltre il primo. Ora: budget
# ampio, ripartito EQUAMENTE tra i file, con troncamenti DICHIARATI al modello.
ATTACH_TOTAL_BUDGET = 120000
ATTACH_MIN_PER_FILE = 6000

# ── Deposito allegati SERVER-SIDE ───────────────────────────
# Il testo estratto vive INTEGRALE sul server (per-utente, TTL 24h): il client
# riceve solo un id. Così la selezione per pertinenza avviene sul file INTERO
# al momento della domanda — nessun taglio cieco prima di conoscerla — e il
# client non rispedisce megabyte di testo a ogni messaggio (App Proxy incluso).
ATTACH_DIR = Path(os.environ.get("APP_DATA_DIR", "/data")) / "attach_cache"
ATTACH_TTL_SECONDS = 24 * 3600


def _attach_user_dir(uid: str) -> Path:
    safe = re.sub(r"[^a-zA-Z0-9._-]", "_", uid or "")[:64] or "u"
    d = ATTACH_DIR / safe
    d.mkdir(parents=True, exist_ok=True)
    return d


def _attach_save(uid: str, text: str, name: str = "") -> str:
    import uuid
    aid = uuid.uuid4().hex
    testa = f"NAME:{(name or '')[:200]}\n" if name and not text.startswith("IMG:") else ""
    (_attach_user_dir(uid) / f"{aid}.txt").write_text(testa + text, encoding="utf-8")
    return aid


def _attach_load(uid: str, aid: str) -> str:
    """Carica il testo di un allegato. L'id è vincolato all'UTENTE della
    sessione: nessun accesso cross-utente, id solo esadecimale.
    Diagnostica LOUD: ogni fallimento viene nominato nei log."""
    import sys
    if not re.fullmatch(r"[0-9a-f]{32}", str(aid or "")):
        print(f"[attach-load] id NON valido: {str(aid)[:40]!r} uid={uid}", file=sys.stderr)
        return ""
    p = _attach_user_dir(uid) / f"{aid}.txt"
    try:
        t = p.read_text(encoding="utf-8")
        if t.startswith("NAME:"):
            t = t.partition("\n")[2]
        return t
    except FileNotFoundError:
        # il file non è sotto QUESTA utenza: esiste altrove? (solo esistenza,
        # mai lettura cross-utente: serve a smascherare un disallineamento uid)
        try:
            altrove = [d.name for d in ATTACH_DIR.iterdir()
                       if d.is_dir() and (d / f"{aid}.txt").is_file()]
        except Exception:
            altrove = []
        print(f"[attach-load] MANCANTE id={aid[:8]} uid={uid} "
              f"{'trovato sotto ALTRA utenza: ' + ','.join(altrove) if altrove else 'assente ovunque'}",
              file=sys.stderr)
        return ""
    except Exception as e:
        print(f"[attach-load] ERRORE lettura id={aid[:8]} uid={uid}: "
              f"{type(e).__name__}: {e}", file=sys.stderr)
        return ""


def _attach_load_image(uid: str, aid: str):
    """Ritorna (media_type, base64) se l'id punta a un'immagine del deposito."""
    t = _attach_load(uid, aid)
    if t.startswith("IMG:"):
        head, _, b64 = t.partition("\n")
        return head[4:] or "image/jpeg", b64
    return None


def _attach_find_by_name(uid: str, name: str) -> str:
    """Risolve un allegato per NOME dentro la cartella dell'UTENTE (mai
    cross-utente): tolleranza per i client vecchi che inviano solo {name}.
    Tra più depositi omonimi vince il più recente."""
    name = (name or "").strip()[:200]
    if not name:
        return ""
    try:
        d = _attach_user_dir(uid)
        for f in sorted(d.glob("*.txt"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                with open(f, encoding="utf-8") as h:
                    prima = h.readline().rstrip("\n")
                if prima == f"NAME:{name}":
                    return f.read_text(encoding="utf-8").partition("\n")[2]
            except Exception:
                continue
    except Exception:
        pass
    return ""


def _attach_purge_old() -> None:
    """Pulizia del deposito: via i testi più vecchi del TTL."""
    import time as _t
    try:
        now = _t.time()
        for f in ATTACH_DIR.rglob("*.txt"):
            try:
                if now - f.stat().st_mtime > ATTACH_TTL_SECONDS:
                    f.unlink()
            except Exception:
                pass
    except Exception:
        pass


def _relevant_slice(text: str, query: str, quota: int):
    """Se il testo supera la quota, seleziona INTESTAZIONI + RIGHE PERTINENTI
    alla domanda (stessa filosofia dello snippet FTS). Pensato per gli Excel:
    'quanti X ci sono?' deve vedere TUTTE le righe di X, e il conteggio viene
    dichiarato nell'intestazione del blocco (contratto dei conteggi).
    Ritorna (testo, per_pertinenza, righe_trovate, righe_incluse, termini)."""
    if len(text) <= quota:
        return text, False, 0, 0, []
    import re as _re
    try:
        from .engines.folder_index import _STOP as _stop
    except Exception:
        _stop = set()
    terms = [w for w in _re.findall(r"[a-zà-ÿ0-9]{3,}", (query or "").lower())
             if w not in _stop]
    lines = text.splitlines()
    head = lines[:15]  # intestazioni: contesto delle colonne
    if terms:
        # GUARDIA anti-termini generici: un termine che matcha oltre metà del
        # file (es. 'asset' in ogni riga di un report asset) non discrimina e
        # inonderebbe la quota: viene scartato dal punteggio.
        low_lines = [ln.lower() for ln in lines[15:]]
        soglia = max(50, len(low_lines) // 2)
        terms = [t for t in terms
                 if sum(1 for ll in low_lines if t in ll) <= soglia]
    if terms:
        scored = []
        for idx, ll in enumerate(low_lines, start=15):
            score = sum(1 for t in terms if t in ll)
            if score > 0:
                scored.append((-score, idx, lines[idx]))
        matched = len(scored)
        scored.sort()
        _marker = "[…righe non pertinenti alla domanda omesse…]"
        picked_idx, used = [], sum(len(l) + 1 for l in head) + len(_marker) + 2
        for _neg, idx, ln in scored:
            if used + len(ln) + 1 > quota:
                break
            picked_idx.append((idx, ln))
            used += len(ln) + 1
        if picked_idx:
            picked_idx.sort()  # ordine originale del file per coerenza
            body = ("\n".join(head) + "\n" + _marker + "\n"
                    + "\n".join(ln for _i, ln in picked_idx))
            return body, True, matched, len(picked_idx), terms
    return text[:quota], False, 0, 0, []


def _build_attach_block(attachments: list, query: str = "", uid: str = "") -> str:
    """Blocco [ALLEGATO ...] per il contesto: budget equo tra i file, selezione
    per pertinenza sui file grandi, CONTEGGI DICHIARATI (calcolati da noi, non
    stimati dal modello) e ogni troncamento esplicitato. È il contratto che
    rende gli allegati affidabili come in un prompt diretto."""
    import sys
    items = []
    falliti = [(str(a.get("name", "allegato")), str(a.get("error")))
               for a in (attachments or [])[:20] if a.get("error")]
    for a in (attachments or [])[:20]:
        if a.get("error"):
            continue
        name = str(a.get("name", "allegato"))
        text = str(a.get("text", "") or "")
        if not text and not a.get("id"):
            # entry {name} da client VECCHIO: il server risolve PER NOME nel
            # deposito dell'utente — tolleranza totale, nessun refresh richiesto
            import sys
            if uid:
                text = _attach_find_by_name(uid, name)
            if text.strip():
                print(f"[attach-ctx] {name}: risolto PER NOME (client vecchio) "
                      f"testo={len(text)} uid={uid}", file=sys.stderr)
            else:
                print(f"[attach-ctx] {name}: ENTRY ANOMALA chiavi={sorted(a.keys())} uid={uid}",
                      file=sys.stderr)
                falliti.append((name, "interfaccia non aggiornata: ricarica la pagina "
                                      "con Ctrl+F5 (o Cmd+Shift+R) e riallega il file"))
                continue
        if not text and a.get("id") and uid:
            aid = str(a.get("id"))
            text = _attach_load(uid, aid)
            if text.startswith("IMG:"):
                text = ""  # le immagini viaggiano come blocchi visione, non testo
            if not text.strip():
                # FAIL LOUDLY: deposito muto = la causa va nominata nei log
                import sys
                _p = _attach_user_dir(uid) / f"{aid}.txt"
                try:
                    _stato = (f"esiste={_p.is_file()}"
                              + (f" size={_p.stat().st_size}" if _p.is_file() else "")
                              + f" dir_file={len(list(_p.parent.glob('*.txt')))}")
                except Exception as _e:
                    _stato = f"stat_err={_e}"
                print(f"[attach-ctx] {name}: DEPOSITO VUOTO id={aid[:8]} uid={uid} "
                      f"{_stato} path={_p}", file=sys.stderr)
        chars_reali = int(a.get("chars") or len(text))
        if text.strip():
            items.append((name, text, chars_reali))
    parts = []
    if falliti:
        # Il modello DEVE spiegare all'utente perché un allegato non c'è,
        # invece di negare l'esistenza di allegati (il motivo era sul chip
        # e moriva lì). Vale anche quando è l'unico "allegato" della domanda.
        dettaglio = "; ".join(f"{n} ({e})" for n, e in falliti)
        parts.append("[ALLEGATI NON CARICATI: " + dettaglio + " — informa "
                     "l'utente che questi file NON sono stati letti e riporta "
                     "il motivo esatto tra parentesi]")
    if not items:
        return "\n\n".join(parts)
    quota = max(ATTACH_MIN_PER_FILE, ATTACH_TOTAL_BUDGET // len(items))
    for i, (name, text, chars_reali) in enumerate(items, 1):
        shown, by_rel, matched, included, terms = _relevant_slice(text, query, quota)
        cut = len(text) > quota
        parziale = chars_reali > len(text)
        header = f"[ALLEGATO {i}/{len(items)}: {name} — {len(text)} caratteri"
        if parziale:
            header += (f" su {chars_reali} ORIGINALI (testo PARZIALE: eventuali "
                       "conteggi vanno dichiarati come parziali)")
        tail = ""
        if by_rel:
            t_str = ", ".join(terms[:6])
            copertura = "TUTTE incluse" if included >= matched else f"incluse le prime {included}"
            header += (f"; RIGHE CORRISPONDENTI ALLA DOMANDA: {matched} "
                       f"(termini: {t_str}), {copertura}]")
            tail = ("\n[selezione per pertinenza: per i conteggi usa il numero "
                    "di righe corrispondenti dichiarato sopra; se un dato manca, "
                    "chiedi all'utente termini più specifici]")
        elif cut:
            header += f"; TRONCATO ai primi {quota} caratteri]"
            tail = "\n[…testo troncato: segnala all'utente che il file è più lungo…]"
        else:
            header += "]"
        parts.append(f"{header}\n{shown}{tail}")
        print(f"[attach-ctx] {name}: testo={len(text)} reali={chars_reali} "
              f"pertinenza={'sì' if by_rel else 'no'} trovate={matched} incluse={included} "
              f"termini={terms[:6]}", file=sys.stderr)
    return "\n\n".join(parts)[:ATTACH_TOTAL_BUDGET + 4000]


def _select_sources(source_links: list, resp_text: str) -> list:
    """Sceglie le fonti da mostrare sotto la risposta.

    Le fonti CITATE nella risposta (per nome) vengono preferite: se il modello
    ha usato solo anagrafica.docx, gli altri risultati Graph non pertinenti non
    vengono mostrati. Report Dynamics e download sono sempre mantenuti. Se
    nessuna fonte risulta citata, si mostrano tutte (fallback prudente).
    Dedup per nome: lo stesso file trovato due volte compare una sola volta."""
    if not source_links:
        return []
    low = (resp_text or "").lower()
    cited, others = [], []
    for s in source_links:
        name = str(s.get("name", "")).strip()
        stem = name.rsplit(".", 1)[0].lower() if "." in name else name.lower()
        if s.get("kind") in ("report", "download"):
            cited.append(s)
        elif name and (name.lower() in low or (len(stem) > 4 and stem in low)):
            # citato col nome completo O senza estensione ("anagrafica")
            cited.append(s)
        else:
            others.append(s)
    # fallback prudente ma non invadente: se il modello non ha citato nessun
    # file, mostra solo i primi 3 (i risultati sono già ordinati per rilevanza)
    chosen = cited if any(s.get("kind") not in ("report", "download") for s in cited) else (cited + others[:3])
    seen, uniq = set(), []
    for s in chosen:
        key = (str(s.get("name", "")).lower() or s.get("url", ""), s.get("kind", ""))
        if key not in seen:
            seen.add(key)
            uniq.append(s)
    return uniq


def _area_settings(base: dict, area_key: str) -> dict:
    """Copia dei settings con il modello Claude dell'area applicato (se
    impostato). Le aree: claude_model_dynamics, claude_model_rewrite,
    claude_model_docgen. Se l'area è vuota resta il modello predefinito."""
    m = (base.get(area_key) or "").strip()
    if not m:
        return base
    s2 = dict(base)
    s2["claude_model"] = m
    return s2


def _client_ip(request: Request) -> str:
    """IP reale del client. Dietro il reverse proxy (Caddy) l'IP diretto è quello
    del proxy: si legge il primo della catena X-Forwarded-For."""
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.headers.get("x-real-ip", "") or (request.client.host if request.client else "")


def _audit(request: Request, username: str, action: str, detail: str = "") -> None:
    """Registra un'attività nell'audit trail con l'IP del client."""
    store.audit_log(username, action, detail, _client_ip(request))


# ── Login / Logout ──────────────────────────────────────────
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if auth.current_user(request):
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(request, "login.html",
                                      {"error": None, **_i18n_ctx(request)})


@app.post("/login", response_class=HTMLResponse)
def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    user = auth.authenticate(username, password)
    if not user:
        _audit(request, (username or "").strip(), "login_fallito", "credenziali non valide o utenza disattivata")
        return templates.TemplateResponse(
            request, "login.html",
            {"error": "Credenziali non valide o utenza disattivata.", **_i18n_ctx(request)},
            status_code=401,
        )
    request.session["user"] = user["username"]
    _audit(request, user["username"], "login", "")
    return RedirectResponse(url="/", status_code=303)


@app.get("/logout")
def logout(request: Request):
    u = request.session.get("user", "")
    if u:
        _audit(request, u, "logout", "")
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


@app.get("/ui-lang")
def set_ui_lang(request: Request):
    """Cambia la lingua dell'interfaccia (it/en). Persistente per utente se loggato."""
    to = i18n.normalize(request.query_params.get("to", "it"))
    nxt = request.query_params.get("next", "/")
    request.session["ui_lang"] = to
    user = auth.current_user(request)
    if user:
        store.set_user_setting(user["username"], "ui_lang", to)
    if not nxt.startswith("/"):
        nxt = "/"
    return RedirectResponse(url=nxt, status_code=303)


# ── Helper impostazioni motore ──────────────────────────────
def admin_settings() -> dict:
    return {
        "claude_api_key": store.get_setting("claude_api_key", ""),
        "claude_model": store.get_setting("claude_model", "claude-opus-4-8"),
        # Modelli per area: il planner Dynamics e la generazione documenti
        # rendono di più con il modello top; la riscrittura query è un compito
        # minuscolo dove il modello economico basta e avanza.
        "claude_model_dynamics": store.get_setting("claude_model_dynamics", "claude-fable-5"),
        "claude_model_rewrite": store.get_setting("claude_model_rewrite", "claude-haiku-4-5-20251001"),
        "claude_model_docgen": store.get_setting("claude_model_docgen", "claude-fable-5"),
        "claude_anonymize": store.get_setting("claude_anonymize", "1") == "1",
        "claude_web_search": store.get_setting("claude_web_search", "0") == "1",
        "cowork_enabled": store.get_setting("cowork_enabled", "0") == "1",
        "lm_url": store.get_setting("lm_url", ""),
        "lm_model": store.get_setting("lm_model", ""),
    }


def anon_names() -> list:
    raw = store.get_setting("anon_dictionary", "")
    return [n.strip() for n in raw.splitlines() if n.strip()]


# ── Chat ────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def chat_page(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    dept = user.get("department") or ""
    uid = user["username"]
    _fold = bool(store.department_folders(dept))
    _od = connectors.is_connected(uid, "onedrive")
    _dy = connectors.is_connected(uid, "dynamics")
    _pb_on = connectors.is_configured("powerbi")  # kill-switch admin incluso
    _pb = _pb_on and connectors.is_connected(uid, "powerbi")
    _avail = {"kb": True, "folder": _fold, "onedrive": _od, "dynamics": _dy, "powerbi": _pb}
    pref_tone = store.get_user_setting(uid, "pref_tone", "Aziendale formale")
    pref_lang = store.get_user_setting(uid, "pref_lang", "Italiano")
    pref_source = store.get_user_setting(uid, "pref_source", "kb")
    if not _avail.get(pref_source, False):
        pref_source = "kb"  # la fonte preferita non è (più) disponibile
    return templates.TemplateResponse(request, "chat.html", _ctx(
        request, user, tones=list(TONES.keys()), langs=list(LANG_INSTR.keys()),
        src_folder_available=_fold,
        src_onedrive_available=_od,
        src_dynamics_available=_dy,
        src_powerbi_available=_pb,
        pbi_enabled=_pb_on,
        pref_engine=store.get_user_setting(uid, "pref_engine", "claude"),
        pref_mode=(lambda _m, _cw: _m if (_m != "cowork" or _cw) else "kb")(store.get_user_setting(uid, "pref_mode", "kb"), store.get_setting("cowork_enabled", "0") == "1"),
        cowork_enabled=store.get_setting("cowork_enabled", "0") == "1",
        pref_tone=pref_tone if pref_tone in TONES else "Aziendale formale",
        pref_lang=pref_lang if pref_lang in LANG_INSTR else "Italiano",
        pref_source=pref_source,
    ))


class ChatRequest(BaseModel):
    mode: str | None = None
    messages: list[dict]
    engine: str = "claude"
    tone: str = "Aziendale formale"
    reply_lang: str = "Italiano"
    free_mode: bool = False
    session_id: str | None = None
    attachments: list[dict] = []
    source: str | None = None  # 'kb' | 'folder' | 'onedrive' | 'dynamics' | 'powerbi' (documentale)


@app.post("/api/chat")
def api_chat(request: Request, body: ChatRequest):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta. Effettua di nuovo l'accesso.")

    def _strip_note_ui(testo: str) -> str:
        """Le note 🔒/🌐 in coda alle risposte sono cornice UI, non conversazione:
        se restano nello storico verso Claude, il modello finisce per fargli eco
        (i '4 lucchetti' del caso Carlos). Via, ovunque compaiano in coda."""
        righe = testo.rstrip().splitlines()
        while righe and (not righe[-1].strip()
                         or re.match(r"^\s*[🔒🌐]\s*\*.*\*\s*$", righe[-1])):
            righe.pop()
        return "\n".join(righe).rstrip()

    clean = []
    for m in body.messages[-20:]:
        role = m.get("role")
        content = (m.get("content") or "").strip()
        if role == "assistant":
            content = _strip_note_ui(content)
        if role in ("user", "assistant") and content:
            clean.append({"role": role, "content": content})
    if not clean or clean[-1]["role"] != "user":
        raise HTTPException(status_code=400, detail="Nessun messaggio utente valido.")

    settings = admin_settings()
    settings["ai_engine"] = "lmstudio" if body.engine == "lmstudio" else "claude"
    settings["tone_key"] = body.tone if body.tone in TONES else "Aziendale formale"
    settings["reply_lang"] = body.reply_lang if body.reply_lang in LANG_INSTR else "Italiano"

    # Recupero conoscenza secondo i TOGGLE dell'utente (Connessioni): conoscenza
    # ChromaDB del dipartimento e/o cartelle del dipartimento. Scoping per area.
    # In MODALITÀ AI LIBERA non si consulta alcuna fonte: risposta da conoscenza generale.
    uid = user["username"]
    _lang = _ui_lang(request, user)
    query = clean[-1]["content"] if clean else ""

    # Allegati della conversazione: hanno priorità e valgono SEMPRE, anche in AI
    # libera (l'utente allega un file e chiede sintesi/elaborazione).
    # ── Modalità ATTIVITÀ (cowork): agente delimitato sulla Conoscenza ──
    if (body.mode or "") == "cowork":
        if store.get_setting("cowork_enabled", "0") != "1":
            def _cw_off():
                yield "data: " + json.dumps({"type": "error", "text":
                    "La modalità Attività non è abilitata: chiedi all'amministratore "
                    "(Admin → Motore)."}, ensure_ascii=False) + "\n\n"
            return StreamingResponse(_cw_off(), media_type="text/event-stream",
                                     headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})
        from . import cowork
        _audit(request, uid, "cowork", f"domanda={query[:160]}")
        dept = user.get("department") or ""
        # Ultimi turni della conversazione: le CORREZIONI ("no, in spagnolo")
        # devono arrivare all'agente col loro contesto, non come compito orfano.
        _conv_rows = []
        for _m in (body.messages or [])[:-1][-8:]:
            _r = "UTENTE" if _m.get("role") == "user" else "ASSISTENTE"
            _c = str(_m.get("content") or "").strip()
            if _c:
                _conv_rows.append(f"{_r}: {_c[:800]}")
        _conv_txt = "\n".join(_conv_rows)[-4000:]
        _note_cw = ""
        try:
            _note_cw = memory.build_note_context(uid)
        except Exception:
            _note_cw = ""
        return StreamingResponse(
            cowork.run(dept, uid, query, _area_settings(settings, "claude_model_dynamics"), anon_names(), conversazione=_conv_txt, note_utente=_note_cw),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    attach_block = _build_attach_block(body.attachments, query, uid)

    # Immagini allegate → blocchi visione Claude (max 5 per domanda)
    images = []
    for a in (body.attachments or [])[:20]:
        if a.get("kind") == "image" and a.get("id"):
            im = _attach_load_image(uid, str(a.get("id")))
            if im:
                images.append(im)
    images = images[:5]
    if images and (body.engine or "claude") != "claude":
        def _img_err():
            yield "data: " + json.dumps({"type": "error", "text":
                "Le immagini allegate sono supportate solo con il motore Claude: "
                "seleziona Claude oppure rimuovi le immagini."}, ensure_ascii=False) + "\n\n"
        return StreamingResponse(_img_err(), media_type="text/event-stream",
                                 headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    # ── FONTE DATI (documentale): UNA sola per domanda, obbligatoria ──
    # Il popup lato client avvisa, ma la validazione vera è qui (client
    # vecchi in cache o chiamate dirette non possono aggirarla).
    src = (body.source or "").strip().lower()
    if not body.free_mode:
        _dept0 = user.get("department") or ""
        _available = {
            "kb": True,
            "folder": bool(store.department_folders(_dept0)),
            "onedrive": connectors.is_connected(uid, "onedrive"),
            "dynamics": connectors.is_connected(uid, "dynamics"),
            "powerbi": connectors.is_configured("powerbi")
                       and connectors.is_connected(uid, "powerbi"),
        }
        _err_msg = ""
        if src not in _available:
            _err_msg = ("Seleziona una fonte dati (Conoscenza, Cartelle, OneDrive, "
                        "Dynamics 365 o Power BI) per la ricerca documentale, oppure "
                        "passa alla modalità AI libera.")
        elif not _available[src]:
            _err_msg = {
                "folder": "Nessuna cartella è configurata per il tuo reparto: chiedi all'amministratore.",
                "onedrive": "OneDrive non è connesso: collegalo dalla pagina Connessioni.",
                "dynamics": "Dynamics 365 non è connesso: collegalo dalla pagina Connessioni.",
                "powerbi": ("Il connettore Power BI è disabilitato dall'amministratore."
                            if not connectors.is_configured("powerbi")
                            else "Power BI non è connesso: collegalo dalla pagina Connessioni."),
            }.get(src, "Fonte dati non disponibile.")
        if _err_msg:
            _msg = _err_msg
            def _src_err():
                yield "data: " + json.dumps({"type": "error", "text": _msg}, ensure_ascii=False) + "\n\n"
            return StreamingResponse(_src_err(), media_type="text/event-stream",
                                     headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    def _build_context() -> tuple[str, list]:
        """Costruisce contesto e fonti secondo i toggle dell'utente. Viene
        eseguito DENTRO il generatore di risposta (in un thread) così il primo
        byte parte subito e la connessione non resta muta durante il retrieval
        (planner Dynamics incluso): i proxy intermedi non vanno in timeout."""
        context = ""
        source_links: list[dict] = []
        if not body.free_mode:
            try:
                dept = user.get("department") or ""
                # Query di ricerca arricchita per TUTTE le fonti (KB, cartelle,
                # OneDrive, Dynamics): i follow-up ("e per il 2025?") ereditano il
                # soggetto dal turno precedente, e la concept map desktop aggiunge
                # il termine documentale ("quanti anni ha X" -> "anagrafica").
                prev_user_q = ""
                for m in reversed(clean[:-1]):
                    if m.get("role") == "user":
                        prev_user_q = m.get("content", "")
                        break
                search_q = knowledge.enrich_query(query, prev_user_q)
                # se ci sono allegati, arricchisci la query con le loro parole chiave
                if attach_block:
                    try:
                        from .engines.onedrive_search import _build_query as _kw
                        akw = _kw(attach_block[:2000])
                        if akw and akw.strip() and akw.strip() not in search_q:
                            search_q = (search_q + " " + akw).strip()
                    except Exception:
                        pass
                # Riscrittura AI della query (opzione admin): il modello riformula
                # la domanda in una query di ricerca con sinonimi e termini inglesi.
                # Costa una piccola chiamata per messaggio; in caso di errore o
                # timeout resta la query deterministica già costruita (mai bloccante).
                # La riscrittura AI (sinonimi bilingue) aiuta SOLO le fonti che
                # cercano in OR o per significato (Conoscenza, Cartelle). Su
                # OneDrive/Graph i termini vengono messi in AND: un sinonimo
                # assente nel documento lo fa sparire dai risultati (verificato
                # nei log: 'biography' -> 0 risultati). Quindi: query riscritta
                # per KB/cartelle, query DETERMINISTICA (stabile e riproducibile)
                # per OneDrive e Dynamics.
                search_q_ai = search_q
                if src in ("kb", "folder") and store.get_setting("search_ai_rewrite", "0") == "1":
                    try:
                        from .orchestrator import complete as _complete
                        rq = _complete(
                            "Trasforma la domanda in una query di ricerca documentale efficace: "
                            "parole chiave essenziali, nomi propri e sinonimi utili, includendo "
                            "gli equivalenti sia in italiano sia in inglese (i documenti sono in "
                            "entrambe le lingue). Massimo 15 parole. Rispondi SOLO con la query, "
                            "senza commenti né punteggiatura finale.",
                            search_q, _area_settings(settings, "claude_model_rewrite"), max_tokens=60, timeout=20).strip()
                        if rq and 2 <= len(rq.split()) <= 25:
                            search_q_ai = rq
                    except Exception as e:
                        import sys
                        print(f"[search] riscrittura AI non disponibile, uso query deterministica: {str(e)[:120]}", file=sys.stderr)
                # FONTE SINGOLA per domanda (flag sotto la chat): la scelta
                # dell'utente decide DOVE cercare; niente ricerche a tappeto.
                parts = []
                if src == "kb":
                    kb_text, kb_names = knowledge.kb_search(dept, search_q_ai, n=5)
                    if kb_text.strip():
                        parts.append(knowledge._fit_budget(
                            ["[Conoscenza dipartimento — " + dept + "]\n" + kb_text], 9500))
                    source_links.extend(knowledge.kb_links(dept, kb_names))
                elif src == "folder":
                    parts.append(knowledge.retrieve(dept, search_q_ai, use_kb=False, use_folder=True))
                elif src == "onedrive":
                    od, od_links = connectors.search_with_links(uid, "onedrive", search_q, max_results=5)
                    if od.strip():
                        parts.append("[OneDrive]\n" + od)
                    source_links.extend(od_links)
                elif src == "dynamics":
                    dy, dy_links = connectors.search_with_links(uid, "dynamics", search_q, max_results=5, current_user_name=uid, ai_settings=settings)
                    if dy.strip():
                        parts.append("[Dynamics 365]\n" + dy)
                    source_links.extend(dy_links)
                elif src == "powerbi":
                    # Query DETERMINISTICA (come Dynamics): il planner Power BI
                    # riceve la domanda arricchita e lavora coi permessi utente.
                    pb, pb_links = connectors.search_with_links(uid, "powerbi", search_q, max_results=5, current_user_name=uid, ai_settings=settings)
                    if pb.strip():
                        parts.append("[Power BI]\n" + pb)
                    source_links.extend(pb_links)
                # budget equo: nessuna fonte (OneDrive/Dynamics incluse) viene
                # scartata in silenzio dal taglio in coda
                # budget alzato (caso Carlos): più documenti per argomento senza
                # tagli in coda; ampiamente dentro la finestra del modello.
                context = knowledge._fit_budget([p for p in parts if p.strip()], 12000)
            except Exception:
                context = ""

        # Gli allegati precedono il resto del contesto (massima priorità).
        # NB: niente taglio cieco del totale — il blocco allegati ha già il suo
        # budget equo e il retrieval il proprio; sommarli è sostenibile.
        if attach_block:
            context = attach_block + ("\n\n" + context if context else "")
        return context, source_links

    # ── Generazione documenti su richiesta (Word/Excel/PPT/PDF) ──
    gen_fmt = docgen.detect_request_with_history(query, clean[:-1])

    # NOTE PERSONALI (Incremento 8): cattura del trigger esplicito
    # ("ricordati che…") PRIMA della risposta, con conferma visibile in chat.
    _nota_conferma = ""
    if memory.note_enabled():
        _nota_testo = memory.estrai_nota(query)
        if _nota_testo:
            _esito = memory.note_add(uid, _nota_testo, origine="chat")
            if _esito.get("ok"):
                _audit(request, uid, "memoria_nota_aggiunta", f"testo={_nota_testo[:120]}")
                _nota_conferma = ("📌 *Aggiunto alle tue note personali: "
                                  f"«{_nota_testo}» — gestibile dalla pagina Connessioni.*\n\n")
            else:
                _nota_conferma = f"📌 *Nota NON salvata: {_esito.get('errore','')}*\n\n"

    # Memoria conversazionale + esempi promossi (solo per la chat normale).
    mem_ctx, fb_ctx = "", ""
    if not gen_fmt:
        try:
            mem_ctx = memory.build_memory_context(uid, exclude_session=body.session_id)
            fb_ctx = memory.build_feedback_context(uid)
        except Exception:
            mem_ctx, fb_ctx = "", ""
    # Le NOTE valgono in chat normale E in generazione documenti.
    _note_ctx = ""
    try:
        _note_ctx = memory.build_note_context(uid)
    except Exception:
        _note_ctx = ""
    if _note_ctx:
        mem_ctx = (_note_ctx + ("\n\n" + mem_ctx if mem_ctx else ""))

    hist_text = ""
    if gen_fmt:
        _turns = []
        for m in clean[:-1][-6:]:
            _role = "Utente" if m.get("role") == "user" else "Assistente"
            _turns.append(f"{_role}: {str(m.get('content', ''))[:800]}")
        hist_text = "\n".join(_turns)

    if gen_fmt:
        _audit(request, uid, "generazione_documento", f"formato={gen_fmt}, fonte={src or '-'}")
    else:
        _audit(request, uid, "chat",
               f"modalita={'libera' if body.free_mode else 'documentale'}, "
               f"fonte={src or '-'}, allegati={len(body.attachments or [])} ({len(attach_block)} car., img={len(images)}), motore={body.engine}")

    _PING = "data: " + json.dumps({"type": "ping"}, ensure_ascii=False) + "\n\n"

    def _in_thread_with_pings(fn):
        """Esegue fn() in un thread e produce un ping SSE ogni 10 secondi finché
        lavora: la connessione non resta mai muta, così gli apparati intermedi
        (VPN, Application Proxy, proxy aziendali) non la chiudono per timeout.
        L'eventuale eccezione del lavoro viene rilanciata nel generatore."""
        res: dict = {}

        def _w():
            try:
                res["v"] = fn()
            except Exception as e:
                res["e"] = e

        th = threading.Thread(target=_w, daemon=True)
        th.start()
        while th.is_alive():
            th.join(timeout=10.0)
            if th.is_alive():
                yield _PING
        if "e" in res:
            raise res["e"]
        yield ("result", res.get("v"))

    def _gen():
        # Primo byte SUBITO: l'apparato di rete vede traffico immediato invece
        # di una connessione in attesa (causa dei 504 delle pagine proxy).
        yield _PING
        # Indicatore di attesa: la ricerca documentale può richiedere tempo
        # (planner Dynamics in testa) — l'utente vede subito cosa sta accadendo.
        if not body.free_mode and not gen_fmt:
            _lbl = {"kb": i18n.t("Conoscenza", _lang), "folder": i18n.t("Cartelle", _lang),
                    "onedrive": "OneDrive", "dynamics": "Dynamics 365",
                    "powerbi": "Power BI"}.get(src, "")
            yield "data: " + json.dumps({"type": "status", "text":
                i18n.t("Ricerca in corso su", _lang) + " " + _lbl + "… " +
                i18n.t("l'operazione può richiedere qualche istante.", _lang)},
                ensure_ascii=False) + "\n\n"
        context, source_links = "", []
        try:
            for item in _in_thread_with_pings(_build_context):
                if isinstance(item, tuple) and item[0] == "result":
                    context, source_links = item[1]
                else:
                    yield item
        except Exception:
            context, source_links = "", []

        if gen_fmt:
            try:
                yield "data: " + json.dumps({"type": "status", "text":
                    i18n.t("Sto preparando il file…", _lang) + " " +
                    i18n.t("l'operazione può richiedere qualche istante.", _lang)},
                    ensure_ascii=False) + "\n\n"
                _tpls = {"docx": None, "pptx": None}
                _tpl_names = {}
                for _f in ("docx", "pptx"):
                    _t = docgen.get_user_template(uid, _f)
                    if _t:
                        _tpls[_f], _tpl_names[_f] = _t[0], _t[1]
                path, fname = None, None
                for item in _in_thread_with_pings(lambda: docgen.generate(gen_fmt, query, context, _area_settings(settings, "claude_model_docgen"), hist_text, templates=_tpls, note_utente=_note_ctx)):
                    if isinstance(item, tuple) and item[0] == "result":
                        path, fname = item[1]
                    else:
                        yield item
                token = connectors.register_download(uid, path, fname)
                _tpl_used = _tpl_names.get("pptx" if gen_fmt == "pptx" else "docx") if gen_fmt in ("docx", "pdf", "pptx") else None
                _msg_gen = (f"Ho preparato **{fname}** sul tuo template **{_tpl_used}**. Puoi scaricarlo qui sotto."
                            if _tpl_used else f"Ho preparato **{fname}**. Puoi scaricarlo qui sotto.")
                yield "data: " + json.dumps({"type": "delta", "text": _msg_gen}, ensure_ascii=False) + "\n\n"
                yield "data: " + json.dumps({"type": "sources", "items": [{"name": fname, "url": "/download/" + token, "kind": "download"}]}, ensure_ascii=False) + "\n\n"
            except Exception as e:
                yield "data: " + json.dumps({"type": "error", "text": "Generazione non riuscita: " + str(e)[:200]}, ensure_ascii=False) + "\n\n"
            return

        # Risposta in streaming (i link NON passano da Claude/anonimizzazione).
        # Accumulo il testo per selezionare poi solo le fonti CITATE.
        resp_parts = []
        if _nota_conferma:
            yield "data: " + json.dumps({"type": "delta", "text": _nota_conferma},
                                        ensure_ascii=False) + "\n\n"
        for chunk in stream_reply(clean, settings, anon_names(), context,
                                  body.free_mode, mem_ctx, fb_ctx, images=images):
            yield chunk
            if '"delta"' in chunk:
                try:
                    evt = json.loads(chunk[6:].strip())
                    if evt.get("type") == "delta":
                        resp_parts.append(evt.get("text", ""))
                except Exception:
                    pass
        uniq = _select_sources(source_links, "".join(resp_parts))
        if uniq:
            yield "data: " + json.dumps({"type": "sources", "items": uniq}, ensure_ascii=False) + "\n\n"

    return StreamingResponse(
        _gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


class PrefsBody(BaseModel):
    engine: str | None = None
    mode: str | None = None
    tone: str | None = None
    lang: str | None = None
    source: str | None = None


@app.get("/admin/logs", response_class=HTMLResponse)
async def admin_logs(request: Request, filtro: str = "", n: int = 400):
    """Log applicativi consultabili dall'app (richiesta di Marco: niente più
    SSH per leggere [attach]/[attach-ctx]/[web]/[folder_index]). Solo admin."""
    user = auth.current_user(request)
    if not user or not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Riservato agli amministratori.")
    righe = diag.tail(n=min(max(n, 10), 800), filtro=filtro[:80])
    corpo = "\n".join(righe) or "(nessuna riga: il buffer si riempie con l'uso)"
    html = ("<html><head><meta charset='utf-8'><meta http-equiv='refresh' content='5'>"
            "<title>ISEOPilot — Log</title><style>body{background:#0a0c10;color:#c9d1d9;"
            "font:12px/1.5 ui-monospace,monospace;padding:16px}h1{color:#EC1D2B;font-size:14px}"
            "form{margin:8px 0}input{background:#12151B;color:#c9d1d9;border:1px solid #262B36;"
            "padding:4px 8px}pre{white-space:pre-wrap;word-break:break-all}</style></head><body>"
            f"<h1>Log applicativi · ultime {len(righe)} righe · refresh 5s</h1>"
            "<form method='get'>filtro: <input name='filtro' value='" + filtro.replace("'", "") + "'>"
            " <input type='submit' value='applica'></form><pre>" + corpo.replace("<", "&lt;")
            + "</pre></body></html>")
    return HTMLResponse(html)


@app.get("/kb/file/{name}")
async def kb_file(request: Request, name: str):
    """Originale di un documento della Conoscenza. Perimetro: l'utente scarica
    SOLO dalla collezione del PROPRIO dipartimento (il percorso è derivato
    dalla sessione, mai dal client); nome sanificato; download in audit."""
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    dept = user.get("department") or ""
    p = knowledge.kb_file_path(dept, name)
    if not p.is_file():
        raise HTTPException(status_code=404, detail=(
            "Originale non archiviato: il documento è stato caricato prima "
            "dell'archiviazione degli originali. Ricaricalo dalla pagina Conoscenza."))
    _audit(request, user["username"], "kb_download", f"file={p.name[:120]}")
    return FileResponse(p, filename=p.name)


@app.post("/api/prefs")
async def api_prefs(request: Request, body: PrefsBody):
    """Preferenze del composer, PER UTENTE: motore, modalità, tono, lingua e
    fonte dati. Salvate a ogni cambio, applicate a ogni apertura della chat
    (su qualunque postazione: vivono lato server, non nel browser)."""
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    uid = user["username"]
    if body.engine in ("claude", "lmstudio"):
        store.set_user_setting(uid, "pref_engine", body.engine)
    if body.mode in ("kb", "free", "cowork"):
        store.set_user_setting(uid, "pref_mode", body.mode)
    if body.tone in TONES:
        store.set_user_setting(uid, "pref_tone", body.tone)
    if body.lang in LANG_INSTR:
        store.set_user_setting(uid, "pref_lang", body.lang)
    if body.source in ("kb", "folder", "onedrive", "dynamics", "powerbi"):
        store.set_user_setting(uid, "pref_source", body.source)
    return {"ok": True}


@app.post("/api/attach")
async def api_attach(request: Request, files: list[UploadFile] = File(...)):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    IMG_EXT = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
    out = []
    for f in files[:20]:
        ext = Path(f.filename or "").suffix.lower()
        if ext in IMG_EXT:
            # IMMAGINI → visione Claude. Ridimensionate e ricodificate JPEG,
            # poi nel deposito per-utente come base64. NB dichiarato in UI:
            # le immagini NON passano dall'anonimizzatore.
            try:
                raw = await f.read()
                if len(raw) > 15 * 1024 * 1024:
                    out.append({"name": f.filename, "ok": False, "error": "Immagine troppo grande (max 15 MB)"})
                    continue
                import base64
                import io as _io
                from PIL import Image as _PILImage
                im = _PILImage.open(_io.BytesIO(raw)); im.load()
                if im.mode != "RGB":
                    bg = _PILImage.new("RGB", im.size, (255, 255, 255))
                    try:
                        bg.paste(im, mask=im.getchannel("A"))
                    except Exception:
                        bg.paste(im.convert("RGB"))
                    im = bg
                lato = max(im.size)
                if lato > 1568:
                    r = 1568 / lato
                    im = im.resize((int(im.width * r), int(im.height * r)), _PILImage.LANCZOS)
                buf = _io.BytesIO(); im.save(buf, "JPEG", quality=85)
                b64 = base64.b64encode(buf.getvalue()).decode()
                aid = _attach_save(user["username"], "IMG:image/jpeg\n" + b64)
                import sys
                print(f"[attach] {user['username']}: {f.filename} -> immagine {im.width}x{im.height}, {len(buf.getvalue())//1024} KB", file=sys.stderr)
                out.append({"id": aid, "name": f.filename, "ok": True, "kind": "image", "chars": 0})
            except Exception as e:
                out.append({"name": f.filename, "ok": False, "error": f"Immagine non leggibile: {str(e)[:80]}"})
            continue
        if ext not in knowledge.ALLOWED_ATTACH_EXT:
            out.append({"name": f.filename, "ok": False,
                        "error": "Tipo non supportato"})
            continue
        try:
            raw = await f.read()
            if len(raw) > 25 * 1024 * 1024:  # 25 MB per file
                out.append({"name": f.filename, "ok": False, "error": "File troppo grande (max 25 MB)"})
                continue
            text = knowledge.extract_attachment_text(f.filename or "", raw)
            import sys
            if not text.strip():
                print(f"[attach] {user['username']}: {f.filename} ({len(raw)} byte) -> NESSUN TESTO", file=sys.stderr)
                out.append({"name": f.filename, "ok": False, "error": "Nessun testo estraibile"})
                continue
            print(f"[attach] {user['username']}: {f.filename} ({len(raw)} byte) -> {len(text)} caratteri estratti", file=sys.stderr)
            # Testo INTEGRALE nel deposito server-side; al client va solo l'id.
            # La selezione per pertinenza avverrà sul testo INTERO quando la
            # domanda esiste — mai più tagli ciechi pre-domanda.
            aid = _attach_save(user["username"], text[:knowledge.ATTACHMENT_MAX_CHARS],
                               name=f.filename or "")
            out.append({"id": aid, "name": f.filename, "ok": True,
                        "chars": len(text)})
        except Exception as e:
            out.append({"name": f.filename, "ok": False, "error": str(e)[:120]})
    return {"attachments": out}


# ── Cronologia conversazioni (per-utente) ──────────────────
def _derive_title(history: list[dict]) -> str:
    for m in history:
        if m.get("role") == "user":
            c = (m.get("content") or "").strip()
            if c and not c.startswith("[CONTESTO"):
                return c[:40] + ("…" if len(c) > 40 else "")
    return "Chat"


class SaveChatReq(BaseModel):
    session_id: str | None = None
    history: list[dict]
    title: str | None = None


class SessionIdReq(BaseModel):
    session_id: str


class RenameReq(BaseModel):
    session_id: str
    title: str


@app.post("/api/chat/save")
def api_chat_save(request: Request, body: SaveChatReq):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    # salva solo se c'è almeno un turno utente reale
    has_user = any(m.get("role") == "user" and (m.get("content") or "").strip()
                   and not (m.get("content") or "").startswith("[CONTESTO") for m in body.history)
    if not has_user:
        return {"session_id": body.session_id or "", "skipped": True}
    title = (body.title or "").strip() or _derive_title(body.history)
    import json as _json
    sid = store.save_chat_session(user["username"], body.session_id or "", title,
                                  _json.dumps(body.history, ensure_ascii=False))
    return {"session_id": sid, "title": title}


@app.get("/api/chat/list")
def api_chat_list(request: Request):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    return {"sessions": store.list_chat_sessions(user["username"])}


@app.get("/api/chat/get")
def api_chat_get(request: Request, sid: str):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    s = store.get_chat_session(user["username"], sid)
    if not s:
        raise HTTPException(status_code=404, detail="Conversazione non trovata.")
    import json as _json
    try:
        hist = _json.loads(s["history"]) if s["history"] else []
    except Exception:
        hist = []
    return {"session_id": s["id"], "title": s["title"], "history": hist}


@app.post("/api/chat/delete")
def api_chat_delete(request: Request, body: SessionIdReq):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    store.delete_chat_session(user["username"], body.session_id)
    return {"ok": True}


@app.post("/api/chat/rename")
def api_chat_rename(request: Request, body: RenameReq):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    title = (body.title or "").strip()[:80] or "Chat"
    store.rename_chat_session(user["username"], body.session_id, title)
    return {"ok": True, "title": title}


# ── Feedback pollice su/giù ────────────────────────────────
class FeedbackReq(BaseModel):
    question: str = ""
    answer: str = ""


@app.post("/api/feedback/good")
def api_feedback_good(request: Request, body: FeedbackReq):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    ans = (body.answer or "").strip()
    if not ans:
        return {"ok": False, "detail": "Nessuna risposta da salvare."}
    store.add_good_answer(user["username"], (body.question or "").strip(), ans)
    return {"ok": True, "promoted": store.count_good_answers(user["username"])}


@app.post("/api/feedback/bad")
def api_feedback_bad(request: Request, body: FeedbackReq):
    user = auth.current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Sessione scaduta.")
    # Segnalazione: non genera apprendimento negativo (come nel desktop). Utile
    # come riscontro all'utente; un'estensione potrebbe registrarla per analisi.
    return {"ok": True}


# ── Impostazioni motore (admin) ─────────────────────────────
@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    return templates.TemplateResponse(request, "admin.html", _ctx(
        request, user,
        claude_model=store.get_setting("claude_model", "claude-opus-4-8"),
        claude_model_dynamics=store.get_setting("claude_model_dynamics", "claude-fable-5"),
        claude_model_rewrite=store.get_setting("claude_model_rewrite", "claude-haiku-4-5-20251001"),
        claude_model_docgen=store.get_setting("claude_model_docgen", "claude-fable-5"),
        claude_key_set=store.has_secret("claude_api_key"),
        claude_anonymize=store.get_setting("claude_anonymize", "1") == "1",
        claude_web_search=store.get_setting("claude_web_search", "0") == "1",
        cowork_enabled=store.get_setting("cowork_enabled", "0") == "1",
        lm_url=store.get_setting("lm_url", ""),
        lm_model=store.get_setting("lm_model", ""),
        anon_dictionary=store.get_setting("anon_dictionary", ""),
        kb_embedding_model=store.get_setting("kb_embedding_model", "all-minilm-l6-v2"),
        kb_reembed_status=store.get_setting("kb_reembed_status", ""),
        search_ai_rewrite=store.get_setting("search_ai_rewrite", "0") == "1",
        kb_mode=store.get_setting("kb_mode", "local"),
        kb_host=store.get_setting("kb_host", "localhost"),
        kb_port=store.get_setting("kb_port", "8000"),
        kb_key_set=store.has_secret("kb_api_key"),
        kb_available=knowledge.kb_available(),
        od_client_id=store.get_setting("od_client_id", DEF_OD_CLIENT_ID),
        od_tenant_id=store.get_setting("od_tenant_id", DEF_OD_TENANT_ID),
        dyn_client_id=store.get_setting("dyn_client_id", DEF_DYN_CLIENT_ID),
        dyn_tenant_id=store.get_setting("dyn_tenant_id", DEF_DYN_TENANT_ID),
        dyn_resource_url=store.get_setting("dyn_resource_url", DEF_DYN_RESOURCE_URL),
        def_od_client_id=DEF_OD_CLIENT_ID, def_od_tenant_id=DEF_OD_TENANT_ID,
        memoria_note_enabled=store.get_setting("memoria_note_enabled", "0") == "1",
        dub_enabled=store.get_setting("dub_enabled", "0") == "1",
        claude_model_dub=store.get_setting("claude_model_dub", "claude-sonnet-4-6"),
        dub_whisper_model=store.get_setting("dub_whisper_model", "small"),
        dub_max_mb=store.get_setting("dub_max_mb", "300"),
        dub_max_min=store.get_setting("dub_max_min", "20"),
        dub_tts_threads=store.get_setting("dub_tts_threads", "2"),
        dub_whisper_threads=store.get_setting("dub_whisper_threads", "4"),
        pbi_enabled=store.get_setting("pbi_enabled", "0") == "1",
        pbi_client_id=store.get_setting("pbi_client_id", DEF_PBI_CLIENT_ID),
        pbi_tenant_id=store.get_setting("pbi_tenant_id", DEF_PBI_TENANT_ID),
        def_dyn_client_id=DEF_DYN_CLIENT_ID, def_dyn_tenant_id=DEF_DYN_TENANT_ID,
        def_dyn_resource_url=DEF_DYN_RESOURCE_URL,
        def_pbi_client_id=DEF_PBI_CLIENT_ID, def_pbi_tenant_id=DEF_PBI_TENANT_ID,
        dyn_catalog=connectors.dyn_catalog_status(),
        dyn_connected=connectors.is_connected(user["username"], "dynamics"),
        dyn_msg=request.query_params.get("dyn_msg", ""),
        dyn_err=request.query_params.get("dyn_err", ""),
        saved=request.query_params.get("saved") == "1",
    ))


@app.post("/admin")
def admin_save(
    request: Request,
    claude_api_key: str = Form(""),
    claude_model: str = Form("claude-opus-4-8"),
    claude_model_dynamics: str = Form("claude-fable-5"),
    claude_model_rewrite: str = Form("claude-haiku-4-5-20251001"),
    claude_model_docgen: str = Form("claude-fable-5"),
    search_ai_rewrite: str = Form("0"),
    claude_anonymize: str = Form("0"),
    claude_web_search: str = Form("0"),
    cowork_enabled: str = Form("0"),
    lm_url: str = Form(""),
    lm_model: str = Form(""),
    anon_dictionary: str = Form(""),
    kb_embedding_model: str = Form("all-minilm-l6-v2"),
    kb_mode: str = Form("local"),
    kb_host: str = Form("localhost"),
    kb_port: str = Form("8000"),
    kb_api_key: str = Form(""),
    od_client_id: str = Form(""),
    od_tenant_id: str = Form(""),
    dyn_client_id: str = Form(""),
    dyn_tenant_id: str = Form(""),
    dyn_resource_url: str = Form(""),
    memoria_note_enabled: str = Form("0"),
    dub_enabled: str = Form("0"),
    claude_model_dub: str = Form(""),
    dub_whisper_model: str = Form(""),
    dub_max_mb: str = Form(""),
    dub_max_min: str = Form(""),
    dub_tts_threads: str = Form(""),
    dub_whisper_threads: str = Form(""),
    pbi_enabled: str = Form("0"),
    pbi_client_id: str = Form(""),
    pbi_tenant_id: str = Form(""),
):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    if claude_api_key.strip():
        store.set_setting("claude_api_key", claude_api_key.strip(), secret=True)
    store.set_setting("claude_model", claude_model.strip() or "claude-opus-4-8")
    store.set_setting("claude_model_dynamics", claude_model_dynamics.strip() or "claude-fable-5")
    store.set_setting("claude_model_rewrite", claude_model_rewrite.strip() or "claude-haiku-4-5-20251001")
    store.set_setting("claude_model_docgen", claude_model_docgen.strip() or "claude-fable-5")
    store.set_setting("search_ai_rewrite", "1" if search_ai_rewrite == "1" else "0")
    store.set_setting("claude_anonymize", "1" if claude_anonymize == "1" else "0")
    store.set_setting("claude_web_search", "1" if claude_web_search == "1" else "0")
    store.set_setting("cowork_enabled", "1" if cowork_enabled == "1" else "0")
    store.set_setting("lm_url", lm_url.strip())
    store.set_setting("lm_model", lm_model.strip())
    store.set_setting("anon_dictionary", anon_dictionary)
    store.set_setting("kb_embedding_model", kb_embedding_model.strip() or "all-minilm-l6-v2")
    store.set_setting("kb_mode", "remote" if kb_mode == "remote" else "local")
    store.set_setting("kb_host", kb_host.strip() or "localhost")
    store.set_setting("kb_port", kb_port.strip() or "8000")
    if kb_api_key.strip():
        store.set_setting("kb_api_key", kb_api_key.strip(), secret=True)
    # Connettori Microsoft (non segreti): se vuoti, ripristina i default ISEO.
    store.set_setting("od_client_id", od_client_id.strip() or DEF_OD_CLIENT_ID)
    store.set_setting("od_tenant_id", od_tenant_id.strip() or DEF_OD_TENANT_ID)
    store.set_setting("dyn_client_id", dyn_client_id.strip() or DEF_DYN_CLIENT_ID)
    store.set_setting("dyn_tenant_id", dyn_tenant_id.strip() or DEF_DYN_TENANT_ID)
    store.set_setting("dyn_resource_url", dyn_resource_url.strip() or DEF_DYN_RESOURCE_URL)
    store.set_setting("memoria_note_enabled", "1" if memoria_note_enabled == "1" else "0")
    store.set_setting("dub_enabled", "1" if dub_enabled == "1" else "0")
    store.set_setting("claude_model_dub", claude_model_dub.strip() or "claude-sonnet-4-6")
    store.set_setting("dub_whisper_model",
                      dub_whisper_model.strip() if dub_whisper_model.strip() in ("small", "medium") else "small")
    for _k, _v, _d in (("dub_max_mb", dub_max_mb, "300"), ("dub_max_min", dub_max_min, "20"),
                       ("dub_tts_threads", dub_tts_threads, "2"),
                       ("dub_whisper_threads", dub_whisper_threads, "4")):
        _n = re.sub(r"[^0-9]", "", _v or "")
        store.set_setting(_k, _n or _d)
    store.set_setting("pbi_enabled", "1" if pbi_enabled == "1" else "0")
    store.set_setting("pbi_client_id", pbi_client_id.strip() or DEF_PBI_CLIENT_ID)
    store.set_setting("pbi_tenant_id", pbi_tenant_id.strip() or DEF_PBI_TENANT_ID)
    return RedirectResponse(url="/admin?saved=1", status_code=303)


# ── Gestione utenti (admin) ─────────────────────────────────
@app.get("/admin/users", response_class=HTMLResponse)
def users_page(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    return templates.TemplateResponse(request, "admin_users.html", _ctx(
        request, user,
        users=store.list_users(), departments=store.list_departments(),
        dub_grants={u["username"]: store.get_user_setting(u["username"], "dub_access", "0") == "1"
                    for u in store.list_users()},
        msg=request.query_params.get("msg", ""), err=request.query_params.get("err", ""),
    ))


@app.post("/admin/users")
def users_create(
    request: Request,
    new_username: str = Form(...),
    new_password: str = Form(...),
    new_department: str = Form(...),
    new_is_admin: str = Form("0"),
):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    uname = new_username.strip()
    _perr = auth.validate_password(new_password)
    if _perr:
        return RedirectResponse(url="/admin/users?err=" + _perr.replace(" ", "+"), status_code=303)
    if not store.department_exists(new_department):
        return RedirectResponse(url="/admin/users?err=Dipartimento+non+valido.", status_code=303)
    ok = store.create_user(uname, auth.hash_password(new_password), new_department,
                           is_admin=(new_is_admin == "1"))
    if not ok:
        return RedirectResponse(url="/admin/users?err=Username+gi%C3%A0+esistente.", status_code=303)
    _audit(request, user["username"], "admin_utente_creato", f"utente={uname}, reparto={new_department}, admin={new_is_admin=='1'}")
    return RedirectResponse(url=f"/admin/users?msg=Utente+{uname}+creato.", status_code=303)


@app.post("/admin/users/update")
def users_update(
    request: Request,
    username: str = Form(...),
    department: str = Form(...),
    is_admin: str = Form("0"),
    active: str = Form("0"),
    dub_access: str = Form("0"),
    reset_password: str = Form(""),
):
    admin = auth.current_user(request)
    if not admin or not admin["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    target = store.get_user(username)
    if not target:
        return RedirectResponse(url="/admin/users?err=Utente+inesistente.", status_code=303)

    want_admin = is_admin == "1"
    want_active = active == "1"
    _dub_prima = store.get_user_setting(username, "dub_access", "0") == "1"
    _dub_dopo = dub_access == "1"
    if _dub_prima != _dub_dopo:
        store.set_user_setting(username, "dub_access", "1" if _dub_dopo else "0")
        _audit(request, admin["username"],
               "dub_grant" if _dub_dopo else "dub_revoca", f"utente={username}")

    # Sicurezza: non lasciare l'app senza alcun amministratore attivo.
    losing_admin = target["is_admin"] and (not want_admin or not want_active)
    if losing_admin and store.admin_count() <= 1:
        return RedirectResponse(
            url="/admin/users?err=Deve+restare+almeno+un+amministratore+attivo.", status_code=303)

    if not store.department_exists(department):
        return RedirectResponse(url="/admin/users?err=Dipartimento+non+valido.", status_code=303)

    pwd_hash = None
    if reset_password.strip():
        _perr = auth.validate_password(reset_password)
        if _perr:
            return RedirectResponse(url="/admin/users?err=" + _perr.replace(" ", "+"), status_code=303)
        pwd_hash = auth.hash_password(reset_password)

    store.update_user(username, department=department, is_admin=want_admin,
                      active=want_active, password_hash=pwd_hash)
    _audit(request, admin["username"], "admin_utente_modificato",
           f"utente={username}, reparto={department}, admin={want_admin}, "
           f"attivo={want_active}, reset_password={bool(reset_password.strip())}")
    return RedirectResponse(url=f"/admin/users?msg=Utente+{username}+aggiornato.", status_code=303)


# ── Audit trail (admin) ─────────────────────────────────────
def _audit_range(preset: str, frm: str, to: str):
    """Calcola (start_iso, end_iso, etichetta) dai filtri temporali."""
    from datetime import datetime, timezone, timedelta
    now = datetime.now(timezone.utc)
    if preset == "24h":
        return (now - timedelta(hours=24)).isoformat(), None, "Ultime 24 ore"
    if preset == "7d":
        return (now - timedelta(days=7)).isoformat(), None, "Ultimi 7 giorni"
    if preset == "30d":
        return (now - timedelta(days=30)).isoformat(), None, "Ultimi 30 giorni"
    if preset == "custom" and (frm or to):
        s = e = None
        try:
            if frm:
                s = datetime.fromisoformat(frm).replace(tzinfo=timezone.utc).isoformat()
            if to:
                e = (datetime.fromisoformat(to).replace(tzinfo=timezone.utc) + timedelta(days=1)).isoformat()
        except Exception:
            pass
        return s, e, f"Dal {frm or '—'} al {to or '—'}"
    return None, None, "Tutto"


@app.get("/admin/audit", response_class=HTMLResponse)
def audit_page(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    qp = request.query_params
    preset = qp.get("preset", "24h")
    frm, to = qp.get("from", ""), qp.get("to", "")
    f_user, f_action = qp.get("user", ""), qp.get("action", "")
    start, end, label = _audit_range(preset, frm, to)
    rows = store.audit_query(start, end, f_user or None, f_action or None, limit=2000)
    return templates.TemplateResponse(request, "admin_audit.html", _ctx(
        request, user, rows=rows, label=label, total=len(rows),
        preset=preset, frm=frm, to=to, f_user=f_user, f_action=f_action,
        actions=store.audit_actions(), usernames=store.audit_usernames(),
    ))


@app.get("/admin/audit/export")
def audit_export(request: Request):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    qp = request.query_params
    start, end, label = _audit_range(qp.get("preset", "24h"), qp.get("from", ""), qp.get("to", ""))
    rows = store.audit_query(start, end, qp.get("user") or None, qp.get("action") or None, limit=100000)

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit"
    headers = ["Data/ora (UTC)", "Utente", "Azione", "Dettaglio", "IP"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = PatternFill("solid", fgColor="EC1D2B")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=(r["ts"] or "").replace("T", " ")[:19])
        ws.cell(row=i, column=2, value=r["username"])
        ws.cell(row=i, column=3, value=r["action"])
        ws.cell(row=i, column=4, value=r["detail"])
        ws.cell(row=i, column=5, value=r["ip"])
    for j, w in enumerate([22, 22, 26, 60, 18], 1):
        ws.column_dimensions[chr(64 + j)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    from datetime import datetime
    fname = f"audit_iseopilot_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    from fastapi.responses import StreamingResponse as _SR
    return _SR(buf,
               media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
               headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Gestione dipartimenti (admin) ───────────────────────────
@app.get("/admin/departments", response_class=HTMLResponse)
def departments_page(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    deps = store.list_departments()
    rows = [{"name": d, "collection": store.collection_for_department(d),
             "folders": store.department_folders(d)} for d in deps]
    return templates.TemplateResponse(request, "admin_departments.html", _ctx(
        request, user, departments=rows,
        msg=request.query_params.get("msg", ""), err=request.query_params.get("err", ""),
    ))


def _list_subdirs(path: str):
    """Elenco sottocartelle di 'path' (solo directory), robusto su share di rete.
    Strumento admin: naviga il filesystem del SERVER per scegliere le cartelle
    ricercabili. Ritorna (path_normalizzato, parent, [sottocartelle])."""
    try:
        base = Path(path or "/").expanduser()
        if not base.is_absolute():
            base = Path("/") / base
        base = base.resolve()
    except Exception:
        base = Path("/")
    subdirs, parent = [], str(base.parent)
    try:
        with os.scandir(str(base)) as it:
            for e in it:
                try:
                    if e.is_dir(follow_symlinks=False) and not e.name.startswith("."):
                        subdirs.append({"name": e.name, "path": str(base / e.name)})
                except OSError:
                    continue
    except (OSError, PermissionError):
        pass
    subdirs.sort(key=lambda d: d["name"].lower())
    return str(base), parent, subdirs


@app.get("/admin/browse", response_class=HTMLResponse)
def admin_browse(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    dept = request.query_params.get("dept", "")
    path = request.query_params.get("path", "/")
    here, parent, subdirs = _list_subdirs(path)
    return templates.TemplateResponse(request, "admin_browse.html", _ctx(
        request, user, dept=dept, here=here, parent=parent, subdirs=subdirs,
        deps=store.list_departments(),
        current_folders=store.department_folders(dept) if dept else [],
    ))


@app.post("/admin/departments/folder/add")
def departments_folder_add(request: Request, department: str = Form(...), path: str = Form(...)):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    if not store.department_exists(department):
        return RedirectResponse(url="/admin/departments?err=Dipartimento+non+valido.", status_code=303)
    p = (path or "").strip()
    if not p or not Path(p).exists():
        return RedirectResponse(
            url=f"/admin/browse?dept={department}&path={p}&err=1", status_code=303)
    store.add_department_folder(department, p)
    return RedirectResponse(
        url=f"/admin/browse?dept={department}&path={p}", status_code=303,
        background=BackgroundTask(knowledge.folder_reindex, p))


@app.post("/admin/departments/folder/remove")
def departments_folder_remove(request: Request, department: str = Form(...), path: str = Form(...)):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    store.remove_department_folder(department, path)
    return RedirectResponse(
        url=f"/admin/departments?msg=Cartella+rimossa+da+{department}.", status_code=303)


@app.post("/admin/departments")
def departments_create(request: Request, new_department: str = Form(...)):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    name = new_department.strip()
    if not name:
        return RedirectResponse(url="/admin/departments?err=Nome+non+valido.", status_code=303)
    ok = store.add_department(name)
    if not ok:
        return RedirectResponse(url="/admin/departments?err=Dipartimento+gi%C3%A0+esistente.", status_code=303)
    return RedirectResponse(url=f"/admin/departments?msg=Dipartimento+{name}+creato.", status_code=303)


# ── Conoscenza: ChromaDB dipartimento + cartella ────────────
@app.get("/knowledge", response_class=HTMLResponse)
def knowledge_page(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    dept = user.get("department") or "—"
    folders = store.department_folders(dept)
    return templates.TemplateResponse(request, "knowledge.html", _ctx(
        request, user,
        kb_available=knowledge.kb_available(),
        kb_collection=store.collection_for_department(dept),
        kb_docs=knowledge.kb_list(dept),
        kb_count=knowledge.kb_count(dept),
        folders=folders,
        folder_count=knowledge.dept_folders_count(dept),
        allowed_ext=", ".join(sorted(knowledge.ALLOWED_UPLOAD_EXT)),
        msg=request.query_params.get("msg", ""), err=request.query_params.get("err", ""),
    ))


@app.post("/knowledge/upload")
def knowledge_upload(request: Request, files: list[UploadFile] = File(...)):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    dept = user.get("department") or ""
    if not knowledge.kb_available():
        return RedirectResponse(url="/knowledge?err=ChromaDB+non+installato+sul+server.", status_code=303)
    ok_count, fail = 0, []
    for f in files:
        if not f or not f.filename:
            continue
        ext = ("." + f.filename.rsplit(".", 1)[-1].lower()) if "." in f.filename else ""
        if ext not in knowledge.ALLOWED_UPLOAD_EXT:
            fail.append(f.filename + " (tipo non supportato)")
            continue
        try:
            raw = f.file.read()
            text = knowledge.extract_text(f.filename, raw)
            done, _msg = knowledge.kb_ingest(dept, f.filename, text)
            if done:
                ok_count += 1
                knowledge.kb_save_file(dept, f.filename, raw)  # originale → link in chat
            else:
                # FAIL LOUDLY: il motivo va mostrato, mai solo il nome del file
                reason = (_msg or "motivo sconosciuto").strip()
                if not (text or "").strip() and ext == ".pdf":
                    reason = ("nessun testo estraibile: probabile PDF scansionato "
                              "(solo immagini, servirebbe OCR)")
                fail.append(f"{f.filename} ({reason})")
                import sys
                print(f"[kb-upload] KO {f.filename}: {reason}", file=sys.stderr)
        except Exception as e:
            fail.append(f"{f.filename} (errore: {str(e)[:120]})")
            import sys
            print(f"[kb-upload] ECCEZIONE {f.filename}: {e}", file=sys.stderr)
    msg = f"{ok_count}+file+indicizzati+nella+collezione+del+dipartimento."
    _audit(request, user["username"], "upload_conoscenza",
           f"reparto={dept}, file_ok={ok_count}, falliti={len(fail)}"
           + (f", motivi={'; '.join(fail[:3])[:300]}" if fail else ""))
    # Modalità AJAX (caricamento a LOTTI dal browser): risposta JSON per lotto.
    # Nata per l'Application Proxy, che rifiuta i POST multipart molto grandi
    # (upload di intere cartelle): il client spezza in richieste piccole.
    if request.query_params.get("ajax") == "1":
        return JSONResponse({"ok": ok_count, "fail": fail})
    if fail:
        elenco = "; ".join(fail)
        if len(elenco) > 800:
            elenco = elenco[:800] + f"… (+{len(fail)} file)"
        return RedirectResponse(
            url="/knowledge?msg=" + msg + "&err=" + ("Non caricati: " + elenco).replace(" ", "+"),
            status_code=303)
    return RedirectResponse(url=f"/knowledge?msg={msg}", status_code=303)


@app.post("/knowledge/delete")
def knowledge_delete(request: Request, filename: str = Form(...)):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    dept = user.get("department") or ""
    ok, _msg = knowledge.kb_delete(dept, filename)
    _audit(request, user["username"], "kb_rimozione_documento", f"file={filename[:120]}, esito={'ok' if ok else 'ko'}")
    if ok:
        knowledge.kb_delete_file(dept, filename)
        return RedirectResponse(url=f"/knowledge?msg=Rimosso+{filename}.", status_code=303)
    # FAIL LOUDLY: il motivo, non un errore generico
    return RedirectResponse(
        url="/knowledge?err=" + ("Errore rimozione: " + (_msg or "motivo sconosciuto")).replace(" ", "+"),
        status_code=303)


@app.post("/knowledge/reindex")
def knowledge_reindex(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    dept = user.get("department") or ""
    if not store.department_folders(dept):
        return RedirectResponse(url="/knowledge?err=Nessuna+cartella+configurata+per+il+dipartimento.", status_code=303)
    ok, msg = knowledge.dept_folders_reindex(dept)
    key = "msg" if ok else "err"
    return RedirectResponse(url=f"/knowledge?{key}=" + msg.replace(" ", "+")[:200], status_code=303)


# ── Connessioni utente ──────────────────────────────────────
# ── Account utente (cambio password self-service) ───────────
@app.get("/account", response_class=HTMLResponse)
def account_page(request: Request, ok: str = "", err: str = ""):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(request, "account.html",
                                      _ctx(request, user, ok=ok, err=err))


@app.post("/account/password")
def account_password(request: Request,
                     current_password: str = Form(""),
                     new_password: str = Form(""),
                     confirm_password: str = Form("")):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    # 1) la password attuale deve essere corretta (anche se la sessione è aperta:
    #    impedisce a chi trova un PC sbloccato di cambiare la password senza saperla)
    if not auth.authenticate(user["username"], current_password):
        return RedirectResponse(url="/account?err=current", status_code=303)
    # 2) validazione della nuova password
    if auth.validate_password(new_password):
        return RedirectResponse(url="/account?err=policy", status_code=303)
    if new_password != confirm_password:
        return RedirectResponse(url="/account?err=match", status_code=303)
    if new_password == current_password:
        return RedirectResponse(url="/account?err=same", status_code=303)
    # 3) salva l'hash scrypt della nuova password (mai in chiaro)
    store.update_user(user["username"], password_hash=auth.hash_password(new_password))
    _audit(request, user["username"], "cambio_password", "")
    return RedirectResponse(url="/account?ok=1", status_code=303)


@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    uname = user["username"]
    dept = user.get("department") or "—"
    return templates.TemplateResponse(request, "user.html", _ctx(
        request, user,
        use_kb=store.get_user_setting(uname, "use_kb", "1") == "1",
        use_folder=store.get_user_setting(uname, "use_folder", "1") == "1",
        use_onedrive=store.get_user_setting(uname, "use_onedrive", "0") == "1",
        use_dynamics=store.get_user_setting(uname, "use_dynamics", "0") == "1",
        use_powerbi=store.get_user_setting(uname, "use_powerbi", "0") == "1",
        kb_count=knowledge.kb_count(dept),
        folders=store.department_folders(dept),
        # Connettori Microsoft: stato reale di connessione per-utente.
        od_connected=connectors.is_connected(uname, "onedrive"),
        dyn_connected=connectors.is_connected(uname, "dynamics"),
        od_configured=connectors.is_configured("onedrive"),
        dyn_configured=connectors.is_configured("dynamics"),
        note_enabled=memory.note_enabled(),
        note_personali=memory.note_list(uname) if memory.note_enabled() else [],
        pbi_enabled=connectors.pbi_enabled(),
        pbi_connected=connectors.is_connected(uname, "powerbi"),
        pbi_configured=connectors.is_configured("powerbi"),
        pbi_catalog=connectors.pbi_catalog_status(uname),
        saved=request.query_params.get("saved") == "1",
    ))


@app.post("/settings")
def settings_save(request: Request, use_kb: str = Form("0"), use_folder: str = Form("0"),
                  use_onedrive: str = Form("0"), use_dynamics: str = Form("0"),
                  use_powerbi: str = Form("0"), ajax: str = Form("")):
    user = auth.current_user(request)
    if not user:
        if ajax == "1":
            return JSONResponse({"ok": False}, status_code=401)
        return RedirectResponse(url="/login", status_code=303)
    uname = user["username"]
    store.set_user_setting(uname, "use_kb", "1" if use_kb == "1" else "0")
    store.set_user_setting(uname, "use_folder", "1" if use_folder == "1" else "0")
    store.set_user_setting(uname, "use_onedrive", "1" if use_onedrive == "1" else "0")
    store.set_user_setting(uname, "use_dynamics", "1" if use_dynamics == "1" else "0")
    store.set_user_setting(uname, "use_powerbi", "1" if use_powerbi == "1" else "0")
    if ajax == "1":
        # Salvataggio automatico (cambio toggle): nessun reload di pagina.
        return JSONResponse({"ok": True})
    return RedirectResponse(url="/settings?saved=1", status_code=303)


@app.post("/connect/{conn}/start")
def connect_start(request: Request, conn: str):
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"ok": False, "error": "Sessione scaduta."}, status_code=401)
    if conn not in connectors.CONNECTORS:
        return JSONResponse({"ok": False, "error": "Connettore non valido."}, status_code=404)
    return JSONResponse(connectors.start(user["username"], conn))


@app.get("/connect/{conn}/poll")
def connect_poll(request: Request, conn: str):
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Sessione scaduta."}, status_code=401)
    if conn not in connectors.CONNECTORS:
        return JSONResponse({"status": "error", "message": "Connettore non valido."}, status_code=404)
    res = connectors.poll_once(user["username"], conn)
    # Appena connesso, attiva automaticamente la fonte: "connesso" implica "in uso".
    if res.get("status") == "connected":
        store.set_user_setting(user["username"], "use_" + conn, "1")
        _audit(request, user["username"], "connettore_connesso", conn)
    return JSONResponse(res)


@app.post("/connect/{conn}/logout")
def connect_logout(request: Request, conn: str):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if conn in connectors.CONNECTORS:
        connectors.disconnect(user["username"], conn)
        _audit(request, user["username"], "connettore_disconnesso", conn)
    return RedirectResponse(url="/settings?saved=1", status_code=303)


@app.post("/knowledge/folder/add")
def knowledge_folder_add(request: Request, path: str = Form(...)):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Solo l'amministratore può aggiungere cartelle.")
    dept = user.get("department") or ""
    p = (path or "").strip()
    if not p or not Path(p).exists():
        return RedirectResponse(url="/knowledge?err=Percorso+non+raggiungibile+dal+server:+" + p.replace(" ", "+")[:160], status_code=303)
    store.add_department_folder(dept, p)
    _audit(request, user["username"], "cartella_aggiunta", f"reparto={dept}, percorso={p}")
    # Indicizzazione automatica in background (anche share di rete): la cartella
    # è subito agganciata; l'indice si costruisce senza bloccare la risposta.
    return RedirectResponse(url="/knowledge?msg=Cartella+aggiunta,+indicizzazione+avviata.",
                            status_code=303, background=BackgroundTask(knowledge.folder_reindex, p))


@app.post("/knowledge/folder/remove")
def knowledge_folder_remove(request: Request, path: str = Form(...)):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Solo l'amministratore può rimuovere cartelle.")
    dept = user.get("department") or ""
    store.remove_department_folder(dept, path)
    _audit(request, user["username"], "cartella_rimossa", f"reparto={dept}, percorso={path}")
    return RedirectResponse(url="/knowledge?msg=Cartella+rimossa.", status_code=303)


@app.post("/admin/kb/reembed")
def admin_kb_reembed(request: Request):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    _audit(request, user["username"], "kb_reindicizzazione", "modello=multilingual")
    knowledge.kb_reembed_async("multilingual")
    return RedirectResponse(
        url="/admin?saved=1", status_code=303)


@app.post("/admin/dynamics/build-catalog")
def admin_build_dyn_catalog(request: Request):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    res = connectors.dyn_build_catalog(user["username"])
    if "errore" in res:
        return RedirectResponse(url="/admin?dyn_err=" + str(res["errore"]).replace(" ", "+")[:200], status_code=303)
    msg = f"Catalogo+generato:+{res.get('count',0)}+entità,+{res.get('relazioni',0)}+relazioni,+{res.get('schema_md',{}).get('file',0)}+schema."
    return RedirectResponse(url="/admin?saved=1&dyn_msg=" + msg, status_code=303)


# ════════════════════════════════════════════════════════════
#  DUB STUDIO (Incremento 9) — gate: kill-switch + grant per-utente
# ════════════════════════════════════════════════════════════
def _dub_gate(request: Request):
    """(user, dubstudio, errore_response). Import LAZY: a kill-switch spento
    il modulo non viene caricato."""
    user = auth.current_user(request)
    if not user:
        return None, None, RedirectResponse(url="/login", status_code=303)
    from . import dubstudio
    if not dubstudio.enabled():
        return None, None, JSONResponse(
            {"error": "Dub Studio non è abilitato."}, status_code=404)
    if not dubstudio.user_allowed(user["username"]):
        return None, None, JSONResponse(
            {"error": "Dub Studio non è abilitato per il tuo account: "
                      "chiedi all'amministratore."}, status_code=403)
    dubstudio._ensure_runner()
    return user, dubstudio, None


@app.get("/dub", response_class=HTMLResponse)
def dub_page(request: Request):
    user, dub, err = _dub_gate(request)
    if err:
        if isinstance(err, RedirectResponse):
            return err
        u = auth.current_user(request)
        raise HTTPException(status_code=err.status_code,
                            detail=json.loads(bytes(err.body))["error"] if u else "")
    from .engines import dub_enroll
    return templates.TemplateResponse(request, "dub.html", _ctx(
        request, user,
        voice=dub.voice_status(user["username"]),
        jobs=dub.job_list(user["username"]),
        langs=dub.LANG_LABELS,
        cfg=dub.settings(),
        worker_ok=dub.worker_alive(),
        enroll_text=dub_enroll.ENROLL_TEXT,
        enroll_text_en=dub_enroll.ENROLL_TEXT_EN,
    ))


@app.post("/dub/voice/check")
async def dub_voice_check(request: Request, file: UploadFile = File(...)):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    raw = await file.read()
    try:
        chk = dub.voice_check(raw, file.filename or "voce.webm")
    except Exception as e:
        return JSONResponse({"ok": False, "errore": f"Audio non leggibile: {e}"})
    return JSONResponse({"ok": chk["ok"], "righe": chk["righe"]})


@app.post("/dub/voice")
async def dub_voice_save(request: Request, file: UploadFile = File(...),
                         consent: str = Form("0")):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    raw = await file.read()
    try:
        esito = dub.voice_save(user["username"], raw, file.filename or "voce.webm",
                               consent == "1")
    except Exception as e:
        return JSONResponse({"ok": False, "errore": f"Salvataggio fallito: {e}"})
    if esito.get("ok"):
        _audit(request, user["username"], "dub_voce_salvata",
               f"durata={dub.voice_status(user['username']).get('duration_s')}s")
    return JSONResponse(esito)


@app.post("/dub/voice/delete")
def dub_voice_delete(request: Request):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    if dub.voice_delete(user["username"]):
        _audit(request, user["username"], "dub_voce_eliminata", "")
    return JSONResponse({"ok": True, "status": dub.voice_status(user["username"])})


@app.post("/dub/job")
async def dub_job_create(request: Request, file: UploadFile = File(...),
                         src: str = Form(...), dst: str = Form(...),
                         mode: str = Form(...)):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    raw = await file.read()
    esito = dub.job_create(user["username"], raw, file.filename or "video.mp4",
                           src, dst, mode)
    if esito.get("ok"):
        _audit(request, user["username"], "dub_job_creato",
               f"id={esito['jid']}, {src}→{dst}, modalità={mode}")
    return JSONResponse(esito)


@app.get("/dub/job/{jid}/status")
def dub_job_status(request: Request, jid: str):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    return JSONResponse(dub.job_status(user["username"], jid))


@app.get("/dub/job/{jid}/review")
def dub_job_review(request: Request, jid: str):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    return JSONResponse(dub.review_get(user["username"], jid))


@app.post("/dub/job/{jid}/review")
async def dub_job_review_save(request: Request, jid: str):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    body = await request.json()
    return JSONResponse(dub.review_save(user["username"], jid,
                                        list(body.get("testi") or [])))


@app.post("/dub/job/{jid}/continue")
def dub_job_continue(request: Request, jid: str):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    esito = dub.job_continue(user["username"], jid)
    if esito.get("ok"):
        _audit(request, user["username"], "dub_job_confermato",
               f"id={jid} → {esito.get('stato')}")
    return JSONResponse(esito)


@app.get("/dub/job/{jid}/srt")
def dub_job_srt(request: Request, jid: str):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    testo = dub.srt_text(user["username"], jid)
    j = dub.job_read(user["username"], jid)
    nome = os.path.splitext(j.get("nome_video", "video"))[0] + f"_{j.get('dst','xx').upper()}.srt"
    return Response(content=testo, media_type="text/plain; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{nome}"'})


@app.get("/dub/job/{jid}/download/{quale}")
def dub_job_download(request: Request, jid: str, quale: str):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    p = dub.output_path(user["username"], jid, quale)
    if not p:
        raise HTTPException(status_code=404, detail="Output non ancora pronto.")
    return FileResponse(str(p), filename=p.name)


@app.post("/dub/job/{jid}/delete")
def dub_job_delete(request: Request, jid: str):
    user, dub, err = _dub_gate(request)
    if err:
        return err
    if dub.job_delete(user["username"], jid):
        _audit(request, user["username"], "dub_job_eliminato", f"id={jid}")
    return JSONResponse({"ok": True})


@app.get("/api/memoria/note")
def memoria_note_list(request: Request):
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"error": "Sessione scaduta."}, status_code=401)
    if not memory.note_enabled():
        return JSONResponse({"enabled": False, "note": []})
    return JSONResponse({"enabled": True, "note": memory.note_list(user["username"])})


@app.post("/api/memoria/note")
def memoria_note_add(request: Request, testo: str = Form(...)):
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"ok": False, "error": "Sessione scaduta."}, status_code=401)
    esito = memory.note_add(user["username"], testo, origine="utente")
    if esito.get("ok") and not esito.get("duplicata"):
        _audit(request, user["username"], "memoria_nota_aggiunta", f"testo={testo[:120]}")
    if not esito.get("ok"):
        return JSONResponse({"ok": False, "error": esito.get("errore", "")})
    return JSONResponse({"ok": True, "note": memory.note_list(user["username"])})


@app.post("/api/memoria/note/delete")
def memoria_note_delete(request: Request, nota_id: str = Form("")):
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"ok": False, "error": "Sessione scaduta."}, status_code=401)
    if nota_id == "all":
        n = memory.note_clear(user["username"])
        _audit(request, user["username"], "memoria_note_svuotate", f"rimosse={n}")
    else:
        if memory.note_delete(user["username"], nota_id):
            _audit(request, user["username"], "memoria_nota_rimossa", f"id={nota_id}")
    return JSONResponse({"ok": True, "note": memory.note_list(user["username"])})


@app.get("/api/template")
def template_status(request: Request):
    """Stato dei template PERSONALI dell'utente (chip nel composer)."""
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"error": "Sessione scaduta."}, status_code=401)
    return JSONResponse(docgen.user_templates_status(user["username"]))


@app.post("/api/template")
async def template_upload(request: Request, file: UploadFile = File(...)):
    """Carica un template personale (.docx/.pptx, no macro): finché presente,
    la generazione documenti lo usa AL POSTO del template di default."""
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"ok": False, "error": "Sessione scaduta."}, status_code=401)
    fname = (file.filename or "").strip()
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
    # anche i MODELLI Office senza macro (.dotx/.potx): normalizzati al salvataggio
    fmt = {"docx": "docx", "dotx": "docx", "pptx": "pptx", "potx": "pptx"}.get(ext, "")
    raw = await file.read()
    if not fmt:
        err = (docgen.validate_office_template(raw, ext or "?", fname)
               or "Sono ammessi solo template .docx (Word) e .pptx (PowerPoint).")
        _audit(request, user["username"], "template_rifiutato", f"file={fname}: {err[:120]}")
        return JSONResponse({"ok": False, "error": err})
    err = docgen.save_user_template(user["username"], fmt, raw, fname)
    if err:
        _audit(request, user["username"], "template_rifiutato", f"file={fname}: {err[:120]}")
        return JSONResponse({"ok": False, "error": err})
    _audit(request, user["username"], "template_caricato", f"formato={fmt}, file={fname}, bytes={len(raw)}")
    return JSONResponse({"ok": True, "status": docgen.user_templates_status(user["username"])})


@app.post("/api/template/delete")
def template_delete(request: Request, fmt: str = Form(...)):
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"ok": False, "error": "Sessione scaduta."}, status_code=401)
    if fmt not in ("docx", "pptx"):
        return JSONResponse({"ok": False, "error": "Formato non valido."}, status_code=400)
    docgen.delete_user_template(user["username"], fmt)
    _audit(request, user["username"], "template_rimosso", f"formato={fmt}")
    return JSONResponse({"ok": True, "status": docgen.user_templates_status(user["username"])})


@app.post("/connect/powerbi/catalog/start")
def pbi_catalog_start(request: Request):
    """Avvia la generazione del catalogo Power BI dell'UTENTE (in background:
    il browser fa polling su /connect/powerbi/catalog/status)."""
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"ok": False, "errore": "Sessione scaduta."}, status_code=401)
    res = connectors.pbi_build_start(user["username"])
    if res.get("ok"):
        _audit(request, user["username"], "powerbi_catalogo", "generazione avviata")
    return JSONResponse(res)


@app.get("/connect/powerbi/catalog/status")
def pbi_catalog_status_route(request: Request):
    user = auth.current_user(request)
    if not user:
        return JSONResponse({"running": False, "errore": "Sessione scaduta."}, status_code=401)
    return JSONResponse(connectors.pbi_build_status(user["username"]))


@app.get("/admin/powerbi/diagnose")
def admin_pbi_diagnose(request: Request):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    return JSONResponse({"report": connectors.pbi_diagnose(user["username"])})


@app.get("/admin/dynamics/diagnose")
def admin_dyn_diagnose(request: Request):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    return JSONResponse({"report": connectors.dyn_diagnose(user["username"])})


@app.get("/admin/dynamics/log")
def admin_dyn_log(request: Request):
    user = auth.current_user(request)
    if not user or not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Accesso riservato all'amministratore.")
    return JSONResponse({"log": connectors.dyn_log_tail(120)})


@app.get("/download/{token}")
def download_file(request: Request, token: str):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    rec = connectors.download_info(user["username"], token)
    if not rec or not Path(rec["path"]).is_file():
        raise HTTPException(status_code=404, detail="File non disponibile o scaduto.")
    from fastapi.responses import FileResponse
    return FileResponse(rec["path"], filename=rec["filename"])


@app.get("/dyn-report/{token}", response_class=HTMLResponse)
def dyn_report(request: Request, token: str):
    user = auth.current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    # Solo il proprietario del report; il path arriva dal registro (non da input
    # utente), quindi niente path traversal. Il report è dati Dynamics riservati.
    path = connectors.report_path(user["username"], token)
    if not path or not Path(path).is_file():
        raise HTTPException(status_code=404, detail="Report non disponibile o scaduto.")
    try:
        html = Path(path).read_text(encoding="utf-8")
    except Exception:
        raise HTTPException(status_code=404, detail="Report non leggibile.")
    return HTMLResponse(content=html)


@app.get("/healthz")
def healthz():
    return {"status": "ok"}
